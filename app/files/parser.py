"""文件解析模块。

这个文件负责把上传文件附件解析成可入库文本。当前优先使用附件 text 字段，
并按文件名/Content-Type 记录解析类型；后续可以扩展 PDF、Excel、Word 的真实解析器。
"""

from __future__ import annotations

import base64
import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from typing import Any
from xml.etree import ElementTree


@dataclass(slots=True)
class ParsedFile:
    # ParsedFile 是文件链路的标准中间结果，后续文件来源不同也统一转成这个结构。
    filename: str
    title: str
    content: str
    parser_type: str
    metadata: dict[str, Any] = field(default_factory=dict)


class UnsupportedFileTypeError(ValueError):
    # 老 Office 二进制格式当前不再返回占位文本，避免把未解析内容伪装成已入库知识。
    pass


class FileParser:
    # FileParser 当前只解析附件里的 text 字段，避免在 API 层直接读任意本地路径造成安全风险。
    def parse(self, attachment: dict[str, Any]) -> ParsedFile | None:
        attachment = _resolve_upload_attachment(attachment)
        filename = attachment.get("filename") or "uploaded-file.txt"
        content_type = attachment.get("content_type")
        parser_type = self._infer_parser_type(filename, content_type)
        if parser_type in {"unsupported_xls", "unsupported_doc"}:
            raise UnsupportedFileTypeError(
                f"unsupported_file_type:{filename}:请转换为 .xlsx、.docx、PDF、CSV 或 TXT 后再上传。"
            )
        text = (attachment.get("text") or "").strip()
        binary = _decode_content_base64(attachment.get("content_base64"))

        parsed_from_binary = self._parse_binary(filename, parser_type, binary) if binary else None
        if parsed_from_binary is not None:
            return parsed_from_binary
        if not text:
            return None

        title = filename.rsplit(".", 1)[0] if "." in filename else filename
        return ParsedFile(
            filename=filename,
            title=title,
            content=text,
            parser_type=parser_type,
            metadata={
                "filename": filename,
                "content_type": content_type,
                "parser_type": parser_type,
                "source": "attachment_text",
            },
        )

    def _parse_binary(self, filename: str, parser_type: str, binary: bytes) -> ParsedFile | None:
        title = filename.rsplit(".", 1)[0] if "." in filename else filename
        if parser_type == "plain_text":
            content = _read_text_bytes(binary).strip()
            if not content:
                return None
            return ParsedFile(
                filename=filename,
                title=title,
                content=content,
                parser_type="plain_text",
                metadata={"filename": filename, "parser_type": "plain_text", "source": "attachment_base64"},
            )
        if parser_type == "csv":
            content, schema = _parse_csv_bytes(binary)
            if not content:
                return None
            return ParsedFile(
                filename=filename,
                title=title,
                content=content,
                parser_type="csv",
                metadata={
                    "filename": filename,
                    "parser_type": "csv",
                    "source": "attachment_base64",
                    "schema_catalog": schema,
                },
            )
        if parser_type == "xlsx":
            content, schema = _parse_xlsx_bytes(binary)
            if not content:
                return None
            return ParsedFile(
                filename=filename,
                title=title,
                content=content,
                parser_type="xlsx",
                metadata={
                    "filename": filename,
                    "parser_type": "xlsx",
                    "source": "attachment_base64",
                    "schema_catalog": schema,
                },
            )
        if parser_type == "pdf":
            content, parser_metadata = _parse_pdf_bytes(binary)
            if not content:
                return None
            return ParsedFile(
                filename=filename,
                title=title,
                content=content,
                parser_type="pdf",
                metadata={
                    "filename": filename,
                    "parser_type": "pdf",
                    "source": "attachment_base64",
                    **parser_metadata,
                },
            )
        if parser_type == "docx":
            content, parser_metadata = _parse_docx_bytes(binary)
            content = content.strip()
            if not content:
                return None
            return ParsedFile(
                filename=filename,
                title=title,
                content=content,
                parser_type="docx",
                metadata={
                    "filename": filename,
                    "parser_type": "docx",
                    "source": "attachment_base64",
                    **parser_metadata,
                },
            )
        return None

    def _infer_parser_type(self, filename: str, content_type: str | None) -> str:
        lowered = filename.lower()
        if lowered.endswith(".pdf") or content_type == "application/pdf":
            return "pdf"
        if lowered.endswith(".csv") or content_type == "text/csv":
            return "csv"
        if lowered.endswith(".xlsx"):
            return "xlsx"
        if lowered.endswith(".xls"):
            return "unsupported_xls"
        if lowered.endswith(".docx"):
            return "docx"
        if lowered.endswith(".doc"):
            return "unsupported_doc"
        return "plain_text"


_file_parser = FileParser()


def get_file_parser() -> FileParser:
    # 当前解析器无外部状态，直接复用模块级单例。
    return _file_parser


def _decode_content_base64(value: str | None) -> bytes | None:
    if not value:
        return None
    try:
        return base64.b64decode(value, validate=True)
    except Exception:
        return None


def _resolve_upload_attachment(attachment: dict[str, Any]) -> dict[str, Any]:
    uri = attachment.get("uri")
    if not isinstance(uri, str) or not uri.startswith("upload://"):
        return attachment
    try:
        from app.uploads.service import get_upload_service

        return get_upload_service().resolve_attachment(attachment)
    except Exception:
        return attachment


def _read_text_bytes(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def _parse_csv_bytes(payload: bytes) -> tuple[str, dict[str, Any]]:
    text = _read_text_bytes(payload)
    rows = list(csv.DictReader(io.StringIO(text)))
    if not rows:
        return text.strip(), {"columns": [], "row_count": 0}
    columns = list(rows[0].keys())
    content_lines = [_flatten_row(row) for row in rows if _flatten_row(row)]
    return "\n".join(content_lines), _build_schema_catalog(columns, rows)


def _parse_xlsx_bytes(payload: bytes) -> tuple[str, dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return "", {"columns": [], "row_count": 0, "error": "openpyxl_missing"}

    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        return "", {"columns": [], "row_count": 0, "error": str(exc)[:200]}
    all_lines: list[str] = []
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = _first_nonempty_row_index(rows)
        if header_index is None:
            continue
        columns = _normalize_headers(rows[header_index])
        records = [
            {
                column: value
                for column, value in zip(columns, row)
                if _has_value(value)
            }
            for row in rows[header_index + 1 :]
        ]
        records = [record for record in records if record]
        for record in records:
            flattened = _flatten_row(record)
            if flattened:
                all_lines.append(f"{sheet.title}: {flattened}")
        sheets.append(
            {
                "sheet_name": sheet.title,
                **_build_schema_catalog(columns, records),
            }
        )
    workbook.close()
    return "\n".join(all_lines), {"sheets": sheets}


def _parse_pdf_bytes(payload: bytes) -> tuple[str, dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError:
        return "", {"error": "pypdf_missing"}

    try:
        reader = PdfReader(io.BytesIO(payload))
    except Exception as exc:
        return "", {"error": str(exc)[:200]}

    page_texts: list[str] = []
    failed_pages: list[dict[str, Any]] = []
    for page_index, page in enumerate(reader.pages, start=1):
        try:
            text = (page.extract_text() or "").strip()
        except Exception as exc:
            failed_pages.append({"page": page_index, "error": str(exc)[:160]})
            continue
        if text:
            page_texts.append(text)
    metadata: dict[str, Any] = {
        "parser_backend": "pypdf",
        "page_count": len(reader.pages),
        "extracted_page_count": len(page_texts),
    }
    if failed_pages:
        metadata["failed_pages"] = failed_pages
    return "\n\n".join(page_texts), metadata


def _parse_docx_bytes(payload: bytes) -> tuple[str, dict[str, Any]]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            xml_payload = archive.read("word/document.xml")
            extra_payloads = [
                archive.read(name)
                for name in sorted(archive.namelist())
                if name.startswith(("word/header", "word/footer")) and name.endswith(".xml")
            ]
    except (KeyError, zipfile.BadZipFile):
        return "", {"error": "bad_docx"}

    content_parts, metadata = _parse_docx_xml(xml_payload)
    for extra_payload in extra_payloads:
        extra_parts, extra_metadata = _parse_docx_xml(extra_payload)
        content_parts.extend(extra_parts)
        metadata["paragraph_count"] += extra_metadata["paragraph_count"]
        metadata["table_count"] += extra_metadata["table_count"]
    metadata["parser_backend"] = "docx_xml"
    return "\n".join(content_parts), metadata


def _parse_docx_xml(xml_payload: bytes) -> tuple[list[str], dict[str, Any]]:
    root = ElementTree.fromstring(xml_payload)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    container = root.find("w:body", ns) or root
    content_parts: list[str] = []
    paragraph_count = 0
    table_count = 0
    for child in list(container):
        if child.tag == _w_tag("p"):
            text = _docx_node_text(child, ns)
            if text:
                paragraph_count += 1
                content_parts.append(text)
        elif child.tag == _w_tag("tbl"):
            rows = _docx_table_rows(child, ns)
            if rows:
                table_count += 1
                content_parts.append("\n".join(rows))
    return content_parts, {"paragraph_count": paragraph_count, "table_count": table_count}


def _docx_table_rows(table: ElementTree.Element, ns: dict[str, str]) -> list[str]:
    rows: list[str] = []
    for row in table.findall("w:tr", ns):
        cells = [_docx_node_text(cell, ns) for cell in row.findall("w:tc", ns)]
        cells = [cell for cell in cells if cell]
        if cells:
            rows.append(" | ".join(cells))
    return rows


def _docx_node_text(node: ElementTree.Element, ns: dict[str, str]) -> str:
    return "".join(text_node.text or "" for text_node in node.findall(".//w:t", ns)).strip()


def _w_tag(name: str) -> str:
    return f"{{http://schemas.openxmlformats.org/wordprocessingml/2006/main}}{name}"


def _build_schema_catalog(columns: list[str], rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "columns": [
            {
                "name": column,
                "data_type": _infer_value_type([row.get(column) for row in rows]),
                "sample_values": _sample_values([row.get(column) for row in rows]),
            }
            for column in columns
        ],
        "row_count": len(rows),
    }


def _first_nonempty_row_index(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows):
        if any(_has_value(value) for value in row):
            return index
    return None


def _normalize_headers(values: tuple[Any, ...]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        name = str(value).strip() if value is not None else ""
        if not name or re.match(r"^Unnamed[:\s]", name, flags=re.I):
            name = f"字段{index}"
        count = used.get(name, 0) + 1
        used[name] = count
        headers.append(name if count == 1 else f"{name}_{count}")
    return headers


def _flatten_row(row: dict[str, Any]) -> str:
    return "；".join(f"{key}: {value}" for key, value in row.items() if _has_value(value))


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    return str(value).strip() not in {"", "-", "nan", "NaN", "None", "null"}


def _sample_values(values: list[Any]) -> list[Any]:
    samples: list[Any] = []
    for value in values:
        if not _has_value(value) or value in samples:
            continue
        samples.append(value)
        if len(samples) >= 3:
            break
    return samples


def _infer_value_type(values: list[Any]) -> str:
    meaningful = [value for value in values if _has_value(value)]
    if not meaningful:
        return "text"
    if all(isinstance(value, bool) for value in meaningful):
        return "boolean"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in meaningful):
        return "integer"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in meaningful):
        return "number"
    return "text"
