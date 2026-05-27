"""本地知识资料读取模块。

这个文件负责把 GustoBot-v2 项目目录内的数据文件读取成统一的 PreparedDocument。
它只做“数据准备”和“文本结构化”：扫描文件、解析文本/表格/JSON、生成稳定 source_id 和 metadata；
真正的切块、embedding、pgvector/内存入库仍然交给 app.kb.service.KnowledgeBaseService。

设计边界：
    1. 这里不直接读取旧项目路径，旧项目资料需要先复制到 v2 的 data/raw/kb。
    2. 这里不连接数据库，也不调用 embedding，方便先 dry-run 检查数据读取结果。
    3. 表格类文件按“每一行一篇文档”展开，便于后续独立追踪来源和增量更新。
"""

from __future__ import annotations

import csv
import json
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping
from xml.etree import ElementTree


@dataclass(slots=True)
class PreparedDocument:
    """数据准备阶段的标准文档对象。

    PreparedDocument 是“文件读取”和“KB 入库”之间的分界线：
    本地文件、CSV 行、Excel 行、JSON 对象最终都会先转成这个结构，再交给 KBService 入库。
    """

    title: str
    content: str
    source_id: str
    metadata: dict[str, Any] = field(default_factory=dict)


class LocalKnowledgeFileLoader:
    """读取本项目内的本地知识资料目录。

    当前支持 txt/md/json/jsonl/csv/xlsx。读取器保持确定性，不做 LLM 改写；
    如果后续要复刻旧项目的“表格行 LLM 重写”，可以在这里增加一个可选 rewrite 步骤。
    """

    SUPPORTED_SUFFIXES = {".txt", ".md", ".markdown", ".json", ".jsonl", ".csv", ".xlsx"}
    SKIPPED_FILENAMES = {"readme.md", ".gitkeep", ".ds_store"}

    def __init__(self, root_dir: str | Path) -> None:
        self.root_dir = Path(root_dir).resolve()

    def load(self) -> list[PreparedDocument]:
        """扫描目录并返回所有可入库文档。

        Returns:
            PreparedDocument 列表。列表顺序按文件路径排序，保证重复运行时 source_id 稳定。
        """
        if not self.root_dir.exists():
            return []
        if not self.root_dir.is_dir():
            raise ValueError(f"数据目录不是文件夹：{self.root_dir}")

        documents: list[PreparedDocument] = []
        for path in sorted(self.root_dir.rglob("*")):
            if not path.is_file() or self._should_skip_file(path):
                continue
            documents.extend(self._load_file(path))
        return documents

    def _should_skip_file(self, path: Path) -> bool:
        # README/.gitkeep 等说明文件不属于业务知识，避免被误入库。
        if path.name.lower() in self.SKIPPED_FILENAMES:
            return True
        return path.suffix.lower() not in self.SUPPORTED_SUFFIXES

    def _load_file(self, path: Path) -> list[PreparedDocument]:
        suffix = path.suffix.lower()
        if suffix in {".txt", ".md", ".markdown"}:
            return [self._load_text_file(path)]
        if suffix == ".csv":
            return self._load_csv_file(path)
        if suffix == ".json":
            return self._load_json_file(path)
        if suffix == ".jsonl":
            return self._load_jsonl_file(path)
        if suffix == ".xlsx":
            return self._load_xlsx_file(path)
        return []

    def _load_text_file(self, path: Path) -> PreparedDocument:
        content = _read_text_with_fallback(path).strip()
        relative_path = self._relative_path(path)
        return PreparedDocument(
            title=path.stem,
            content=content,
            source_id=f"local:{relative_path}",
            metadata=self._base_metadata(path, parser_type="plain_text"),
        )

    def _load_csv_file(self, path: Path) -> list[PreparedDocument]:
        text = _read_text_with_fallback(path)
        rows = csv.DictReader(text.splitlines())
        documents: list[PreparedDocument] = []
        for row_number, row in enumerate(rows, start=2):
            document = self._document_from_mapping(
                path,
                row,
                parser_type="csv_row",
                suffix=f"row:{row_number}",
                title_suffix=f"第 {row_number} 行",
                extra_metadata={"row_number": row_number},
            )
            if document:
                documents.append(document)
        return documents

    def _load_json_file(self, path: Path) -> list[PreparedDocument]:
        payload = json.loads(_read_text_with_fallback(path))
        if isinstance(payload, dict) and isinstance(payload.get("documents"), list):
            items = payload["documents"]
        elif isinstance(payload, list):
            items = payload
        else:
            items = [payload]
        return self._load_json_items(path, items, parser_type="json")

    def _load_jsonl_file(self, path: Path) -> list[PreparedDocument]:
        items: list[Any] = []
        for line in _read_text_with_fallback(path).splitlines():
            stripped = line.strip()
            if stripped:
                items.append(json.loads(stripped))
        return self._load_json_items(path, items, parser_type="jsonl")

    def _load_xlsx_file(self, path: Path) -> list[PreparedDocument]:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            # 当前运行环境可能还没安装 openpyxl；为了让数据准备 dry-run 不被依赖阻断，
            # 这里用标准库读取 xlsx 内部 XML，支持常见的纯文本/数字表格。
            return self._load_xlsx_file_with_stdlib(path)

        workbook = load_workbook(path, read_only=True, data_only=True)
        documents: list[PreparedDocument] = []
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row_number, headers = _read_header_row(rows)
            if not headers:
                continue

            for row_number, values in enumerate(rows, start=header_row_number + 1):
                row = _row_to_mapping(headers, values)
                document = self._document_from_mapping(
                    path,
                    row,
                    parser_type="xlsx_row",
                    suffix=f"{sheet.title}:row:{row_number}",
                    title_suffix=f"{sheet.title} 第 {row_number} 行",
                    extra_metadata={"sheet_name": sheet.title, "row_number": row_number},
                )
                if document:
                    documents.append(document)
        workbook.close()
        return documents

    def _load_xlsx_file_with_stdlib(self, path: Path) -> list[PreparedDocument]:
        sheets = _read_xlsx_sheets_with_stdlib(path)
        documents: list[PreparedDocument] = []
        for sheet_name, sheet_rows in sheets.items():
            if not sheet_rows:
                continue
            header_row_number, headers = _read_header_row(iter(sheet_rows))
            if not headers:
                continue

            for row_number, values in enumerate(sheet_rows[header_row_number:], start=header_row_number + 1):
                row = _row_to_mapping(headers, tuple(values))
                document = self._document_from_mapping(
                    path,
                    row,
                    parser_type="xlsx_row_stdlib",
                    suffix=f"{sheet_name}:row:{row_number}",
                    title_suffix=f"{sheet_name} 第 {row_number} 行",
                    extra_metadata={"sheet_name": sheet_name, "row_number": row_number},
                )
                if document:
                    documents.append(document)
        return documents

    def _load_json_items(
        self,
        path: Path,
        items: Iterable[Any],
        *,
        parser_type: str,
    ) -> list[PreparedDocument]:
        documents: list[PreparedDocument] = []
        for index, item in enumerate(items, start=1):
            if isinstance(item, Mapping):
                document = self._document_from_mapping(
                    path,
                    item,
                    parser_type=parser_type,
                    suffix=f"item:{index}",
                    title_suffix=f"第 {index} 条",
                    extra_metadata={"item_index": index},
                )
            else:
                text = str(item).strip()
                document = self._document_from_text(
                    path,
                    text,
                    parser_type=parser_type,
                    suffix=f"item:{index}",
                    title_suffix=f"第 {index} 条",
                    extra_metadata={"item_index": index},
                )
            if document:
                documents.append(document)
        return documents

    def _document_from_mapping(
        self,
        path: Path,
        row: Mapping[str, Any],
        *,
        parser_type: str,
        suffix: str,
        title_suffix: str,
        extra_metadata: dict[str, Any],
    ) -> PreparedDocument | None:
        clean_row = {str(key): _json_safe_value(value) for key, value in row.items() if key is not None}
        content = _first_nonempty(clean_row, ("content", "text", "body", "rewritten_content")) or _flatten_mapping(clean_row)
        return self._document_from_text(
            path,
            content,
            parser_type=parser_type,
            suffix=suffix,
            title_suffix=title_suffix,
            extra_metadata={**extra_metadata, "row_data": clean_row},
        )

    def _document_from_text(
        self,
        path: Path,
        content: str,
        *,
        parser_type: str,
        suffix: str,
        title_suffix: str,
        extra_metadata: dict[str, Any],
    ) -> PreparedDocument | None:
        text = (content or "").strip()
        if not text:
            return None

        relative_path = self._relative_path(path)
        title = f"{path.stem} {title_suffix}".strip()
        return PreparedDocument(
            title=title,
            content=text,
            source_id=f"local:{relative_path}#{suffix}",
            metadata={
                **self._base_metadata(path, parser_type=parser_type),
                **extra_metadata,
            },
        )

    def _base_metadata(self, path: Path, *, parser_type: str) -> dict[str, Any]:
        return {
            "filename": path.name,
            "relative_path": self._relative_path(path),
            "parser_type": parser_type,
            "source": "local_file",
            "suffix": path.suffix.lower(),
        }

    def _relative_path(self, path: Path) -> str:
        # source_id 统一使用正斜杠，避免 Windows 反斜杠影响跨平台日志和引用。
        return path.resolve().relative_to(self.root_dir).as_posix()


def _read_text_with_fallback(path: Path) -> str:
    """用常见中文编码读取文本文件。

    旧项目和人工整理资料可能混用 utf-8-sig、utf-8、gb18030；
    这里集中做容错，避免数据准备脚本因为单个文件编码差异直接失败。
    """

    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _read_header_row(rows: Iterable[tuple[Any, ...]]) -> tuple[int, list[str]]:
    for row_number, values in enumerate(rows, start=1):
        if any(_has_value(value) for value in values):
            return row_number, _normalize_headers(values)
    return 0, []


def _normalize_headers(values: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        name = str(value).strip() if value is not None else ""
        if not name or re.match(r"^Unnamed[:\s]", name, flags=re.I):
            name = f"字段{index}"
        name = name.replace("（", "(").replace("）", ")")
        count = used.get(name, 0) + 1
        used[name] = count
        headers.append(name if count == 1 else f"{name}_{count}")
    return headers


def _row_to_mapping(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: value
        for header, value in zip(headers, values)
        if _has_value(value)
    }


def _flatten_mapping(row: Mapping[str, Any]) -> str:
    pairs: list[str] = []
    for key, value in row.items():
        if not _has_value(value):
            continue
        key_text = str(key).strip()
        if not key_text or re.match(r"^Unnamed[:\s]", key_text, flags=re.I):
            continue
        pairs.append(f"{key_text}: {value}")
    return "\n".join(pairs)


def _first_nonempty(row: Mapping[str, Any], keys: tuple[str, ...]) -> str | None:
    lowered = {str(key).lower(): value for key, value in row.items()}
    for key in keys:
        value = lowered.get(key)
        if _has_value(value):
            return str(value)
    return None


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text not in {"", "-", "nan", "NaN", "None", "null"}


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    return value


def _read_xlsx_sheets_with_stdlib(path: Path) -> dict[str, list[list[Any]]]:
    """使用标准库读取简单 xlsx 工作簿。

    xlsx 本质是一个 zip 包，里面存放 workbook、sharedStrings 和 worksheet XML。
    这个 fallback 不追求完整 Excel 兼容，只覆盖数据准备常见场景：第一行表头，后续行为文本/数字单元格。
    """

    with zipfile.ZipFile(path) as archive:
        shared_strings = _read_shared_strings(archive)
        sheet_paths = _read_workbook_sheet_paths(archive)
        return {
            sheet_name: _read_sheet_rows(archive, sheet_path, shared_strings)
            for sheet_name, sheet_path in sheet_paths.items()
        }


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        payload = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ElementTree.fromstring(payload)
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings: list[str] = []
    for item in root.findall("main:si", ns):
        parts = [node.text or "" for node in item.findall(".//main:t", ns)]
        strings.append("".join(parts))
    return strings


def _read_workbook_sheet_paths(archive: zipfile.ZipFile) -> dict[str, str]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("pkgrel:Relationship", ns)
        if "Id" in rel.attrib and "Target" in rel.attrib
    }
    sheet_paths: dict[str, str] = {}
    for sheet in workbook.findall(".//main:sheet", ns):
        name = sheet.attrib.get("name", "Sheet")
        rel_id = sheet.attrib.get(f"{{{ns['rel']}}}id")
        target = rel_targets.get(rel_id or "")
        if not target:
            continue
        normalized_target = target.lstrip("/")
        sheet_paths[name] = normalized_target if normalized_target.startswith("xl/") else f"xl/{normalized_target}"
    return sheet_paths


def _read_sheet_rows(
    archive: zipfile.ZipFile,
    sheet_path: str,
    shared_strings: list[str],
) -> list[list[Any]]:
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ElementTree.fromstring(archive.read(sheet_path))
    rows: list[list[Any]] = []
    for row in root.findall(".//main:sheetData/main:row", ns):
        values: list[Any] = []
        for cell in row.findall("main:c", ns):
            index = _cell_column_index(cell.attrib.get("r", ""))
            while len(values) < index:
                values.append(None)
            values.append(_read_cell_value(cell, shared_strings, ns))
        rows.append(values)
    return rows


def _read_cell_value(
    cell: ElementTree.Element,
    shared_strings: list[str],
    ns: dict[str, str],
) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//main:t", ns))
    value_node = cell.find("main:v", ns)
    if value_node is None or value_node.text is None:
        return None
    value = value_node.text
    if cell_type == "s":
        index = int(value)
        return shared_strings[index] if 0 <= index < len(shared_strings) else value
    return value


def _cell_column_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Za-z]", "", cell_ref).upper()
    if not letters:
        return 0
    index = 0
    for letter in letters:
        index = index * 26 + (ord(letter) - ord("A") + 1)
    return index - 1
